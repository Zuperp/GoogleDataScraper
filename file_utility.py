# file_utils.py

import pandas as pd
from openpyxl import load_workbook
import os

def detect_header_row_and_columns(excel_file, search_cols=("Keyword", "HITS"), search_rows=8):
    """
    Returns (header_row, found_cols)
    found_cols: dict mapping search_col names to their index
    """
    df = pd.read_excel(excel_file, header=None, nrows=search_rows)
    header_row = None
    found_cols = {}
    for i, row in df.iterrows():
        for col_idx, value in enumerate(row):
            if isinstance(value, str):
                for search in search_cols:
                    if value.strip().lower() == search.lower():
                        found_cols[search] = col_idx
        if all(col in found_cols for col in search_cols):
            header_row = i
            break
    return header_row, found_cols

def read_keywords(excel_file, keyword_col, header_row, max_empty=2):
    """
    Returns a list of (df_idx, keyword)
    Only returns keywords in the main block (stops at max_empty blank rows)
    """
    df = pd.read_excel(excel_file, header=header_row)
    keywords = []
    empty_count = 0
    for idx, val in enumerate(df[keyword_col]):
        if pd.notnull(val) and str(val).strip():
            keywords.append((idx, str(val).strip()))
            empty_count = 0
        else:
            empty_count += 1
            if empty_count > max_empty:
                break
    return keywords

def read_mock_hits(mock_file):
    """
    Reads the first column of the given Excel file as a list.
    """
    return pd.read_excel(mock_file, header=None).iloc[:, 0].tolist()

def update_hits_column(
    excel_file, indices_to_update, hits_list, header_row, hits_col_idx, overwrite=True, save_as=None
):
    """
    Updates only the HITS column of the given Excel file.
    indices_to_update: list of DataFrame indices to update
    hits_list: list of integers to write (same length)
    If overwrite: saves to excel_file, else saves to save_as
    """
    wb = load_workbook(excel_file)
    ws = wb.active

    excel_hits_col_idx = hits_col_idx + 1  # openpyxl is 1-indexed
    start_data_row = header_row + 2  # Data starts 2 rows after header (header is 0-based, Excel is 1-based)

    for list_idx, df_idx in enumerate(indices_to_update):
        excel_row = start_data_row + df_idx
        ws.cell(row=excel_row, column=excel_hits_col_idx, value=hits_list[list_idx])

    output_file = excel_file if overwrite else (save_as or "output.xlsx")
    wb.save(output_file)
    return output_file
